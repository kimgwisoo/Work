# -*- coding utf-8 -*-
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border, Side, Alignment
from datetime import date, timedelta
from openpyxl.styles import PatternFill
import pymssql

#DB 연결부분
def dbConnection () :
    global conn
    try :
        conn = pymssql.connect(host="192.168.100.12", user='MCF', password='MCF_!@#$', database='MCF', charset='utf8')
        print("DB Connect Success!")
    except Exception as e:    
        print("DB Connect 실패!")

#conn = pymssql.connect(host="127.0.0.1", user='tester', password='1234', database='AdventureWorks2014')
dbConnection()
cursor = conn.cursor() 

def insert_value(ws = openpyxl.workbook.workbook.Workbook, sql_exe = None):
    cursor.execute(sql_exe)
    row = cursor.fetchone()
    count = 1
    while(row):
        count += 1
        
        cell_A = "A{}".format(count)
        ws[cell_A].border = boxA
        ws[cell_A].alignment = center_ali
        ws[cell_A] = count-1        
        row_index = 0
        
        for col in cell_range:
            cell_da = "{}{}".format(col,count)
            ws[cell_da].border = boxA            
            ws[cell_da].alignment = center_ali
            ws[cell_da] = row[row_index]
            row_index += 1
            if col is 'C':
                ws[cell_da].alignment = left_ali
        
        row = cursor.fetchone()        
    return count - 1
 

#cell style 정의
font_A = Font(name = '맑은 고딕', bold='True')
boxA = Border(top = Side(style='thin'), left = Side(style='thin'), right = Side(style='thin'),bottom = Side(style='thin'))
center_ali = Alignment(horizontal='center',vertical='center')
left_ali = Alignment(horizontal='general',vertical='center')
gray_fill = PatternFill('solid',fgColor='d8d8d8')

#날짜 변수 선언
today = date.today()
yesterday = (date.today() - timedelta(1))
week_day = (date.today() - timedelta(7))
day_check = today.weekday()

start_time = "05:00"
end_time = "05:00"
#범위
data_time = "{} {} ~ {} {}".format(yesterday.strftime('%m.%d'), start_time, today.strftime('%m.%d'), end_time)
data_time_week = f"{week_day.strftime('%m.%d')} {start_time} ~ {today.strftime('%m.%d')} {end_time}"

#시트 생성
wb = openpyxl.Workbook()
if day_check == 4:
    ws1 = wb.create_sheet('탐지건수', 0)
    ws2 = wb.create_sheet('경유지(05시~05시)', 1)
    ws3 = wb.create_sheet('유포지(05시~05시)', 2)
    ws4 = wb.create_sheet('경유지(전주05시~금주05시)', 3)
    ws5 = wb['Sheet']
    ws3.title = '유포지(전주05시~금주05시)'
else:
    ws1 = wb.create_sheet('탐지건수', 0)
    ws2 = wb.create_sheet('경유지(05시~05시)', 1)
    ws3 = wb['Sheet']
    ws3.title = '유포지(05시~05시)'
#SQL 실행

excute_order ="SELECT b.create_dt '탐지일시' , b.relay_url '경유지', (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END ) '업체명' FROM    [MCF].[DBO].RELAY_URL_HISTORY B, [MCF].[DBO].DOMAIN_MASTER C WHERE  b.relay_domain = c.domain and b.create_dt between '{} 05:00:00' and '{} 05:00:00' order by b.create_dt;".format(yesterday.strftime('%Y/%m/%d') ,today.strftime('%Y/%m/%d'))
cursor.execute(excute_order)

row = cursor.fetchone()

cell_range = ['B','C','D']
first_row = ['A','B','C','D']
#텍스트 파일 생성 및 excel 값 삽입
relay_name = "{}_경유지.txt".format(today.strftime('%y%m%d'))

with open(relay_name,'w',encoding = 'utf-8') as fw:
    count = 1
    while(row):
        count += 1
        fw.write("{}\n".format(row[1]))
        cell_A = "A{}".format(count)
        ws2[cell_A].border = boxA
        ws2[cell_A].alignment = center_ali
        ws2[cell_A] = count-1        
        row_index = 0
        
        for col in cell_range:
            cell_da = "{}{}".format(col,count)
            ws2[cell_da].border = boxA            
            ws2[cell_da].alignment = center_ali
            ws2[cell_da] = row[row_index]
            row_index += 1
            if col is 'C':
                ws2[cell_da].alignment = left_ali
        
        row = cursor.fetchone()        
    count_relay_url = count - 1

excute_order = "SELECT d.create_dt '탐지일시', d.diffusion_url '유포지' ,  (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END )  '업체명' FROM    [MCF].[DBO].DIFFUSION_URL_LIST B,  [MCF].[DBO].DOMAIN_MASTER C, [MCF].[dbo].[DIFFUSION_ANALYSIS_GROUP_LIST] D WHERE  b.diffusion_domain = c.domain and b.diffusion_url_hash = d.diffusion_url_hash and     d.create_dt between  '{} 05:00:00'  and '{} 05:00:00' order by d.create_dt;".format(yesterday.strftime('%Y/%m/%d') ,today.strftime('%Y/%m/%d'))

cursor.execute(excute_order)
row = cursor.fetchone()

diffusion_name = "{}_유포지.txt".format(today.strftime('%y%m%d'))
with open(diffusion_name,'w',encoding = 'utf-8') as fw:
    count = 1
    while(row):
        count += 1
        fw.write("{}\n".format(row[1]))
        cell_A = "A{}".format(count)
        ws3[cell_A].border = boxA
        ws3[cell_A].alignment = center_ali
        ws3[cell_A] = count-1        
        row_index = 0
        
        for col in cell_range:
            cell_da = "{}{}".format(col,count)
            print(cell_da)
            ws3[cell_da].border = boxA            
            ws3[cell_da].alignment = center_ali
            ws3[cell_da] = row[row_index]
            row_index += 1
            if col is 'C':
                ws3[cell_da].alignment = left_ali
        
        row = cursor.fetchone()        
    count_diffusion_url = count - 1






if day_check == 4:
    excute_order = f"SELECT b.create_dt '탐지일시' , b.relay_url '경유지', (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END ) '업체명' FROM    [MCF].[DBO].RELAY_URL_HISTORY B, [MCF].[DBO].DOMAIN_MASTER C WHERE  b.relay_domain = c.domain and b.create_dt between '{week_day.strftime('%Y/%m/%d')} {start_time}:00' and '{today.strftime('%Y/%m/%d')} {end_time}:00' order by b.create_dt;"
    count_week_relay = insert_value(ws4, excute_order)
    
    excute_order = "SELECT d.create_dt '탐지일시', d.diffusion_url '유포지' ,  (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END )  '업체명' FROM    [MCF].[DBO].DIFFUSION_URL_LIST B,  [MCF].[DBO].DOMAIN_MASTER C, [MCF].[dbo].[DIFFUSION_ANALYSIS_GROUP_LIST] D WHERE  b.diffusion_domain = c.domain and b.diffusion_url_hash = d.diffusion_url_hash and     d.create_dt between  '{} {}:00'  and '{} {}:00' order by d.create_dt;".format(week_day.strftime('%Y/%m/%d'),start_time ,today.strftime('%Y/%m/%d'),end_time)    
    count_week_dif = insert_value(ws5, excute_order)
'''
###################
#####탐지 건수######
###################
'''


#A열

ws1['A2'] = '※ 탐지 건수'
ws1['A2'].font = font_A

ws1['A3'] = '일시'
ws1['A4'] =  data_time
ws1['A7'] = '※ 보고기준'
ws1['A7'].font = font_A

ws1['A8'] = ' - 16시 보고 : 전일 16시 ~ 금일 16시'
ws1['A9'] = '                   금일 05시 ~ 금일 16시'
ws1['A10'] = ' - 05시 보고 : 전일 05시 ~ 금일 05시'




#B열
ws1.column_dimensions['A'].width = 35.2 #픽셀 수 = 간격
ws1['B3'] = '경유지'
ws1['B4'] = count_relay_url

ws1['C3'] = '유포지'
ws1['C4'] = count_diffusion_url

ws1['D3'] = '합계'
ws1['D4'] = count_relay_url + count_diffusion_url

###금요일일때####
if day_check == 4:
    ws1['A5'] = week_day
    
    ws1['B5'] = count_week_relay
    ws1['C5'] = count_week_dif
    ws1['D5'] = count_week_relay + count_week_dif
    for i in frist_row:
        cell_op = f"{i}5"
        ws1[cell_op].border = boxA
        ws1[cell_op].alignment = center_ali
    


for i in first_row:
    for x in range(3,5):
        cell_da = "{}{}".format(i,x)
        ws1[cell_da].border = boxA
        ws1[cell_da].alignment = center_ali
        

'''
#########################
####경유지(05시~05시)#####
#########################
'''

ws2['A1'] = '번호'
ws2['B1'] = '탐지일'
ws2['C1'] = 'URL'
ws2['D1'] = '업체명'

ws2.column_dimensions['A'].width = 9 #픽셀 수 = 간격
ws2.column_dimensions['B'].width = 25.4
ws2.column_dimensions['C'].width = 84.1
ws2.column_dimensions['D'].width = 28.4

for i in first_row:
    cell_da = "{}{}".format(i,1)
    ws2[cell_da].border = boxA
    ws2[cell_da].fill = gray_fill
    ws2[cell_da].alignment = center_ali



'''
###########################
#####유포지(05시~05시)######
###########################
'''

ws3['A1'] = '번호'
ws3['B1'] = '탐지일'
ws3['C1'] = 'URL'
ws3['D1'] = '업체명'

ws3.column_dimensions['A'].width = 9 #픽셀 수 = 간격
ws3.column_dimensions['B'].width = 25.4
ws3.column_dimensions['C'].width = 84.1
ws3.column_dimensions['D'].width = 28.4

for i in first_row:
    cell_da = "{}{}".format(i,1)
    ws3[cell_da].border = boxA
    ws3[cell_da].fill = gray_fill
    ws3[cell_da].alignment = center_ali

####금요일일때 excel 틀 ##
if day_check == 4:
    ws4['A1'] = '번호'
    ws4['B1'] = '탐지일'
    ws4['C1'] = 'URL'
    ws4['D1'] = '업체명'

    ws4.column_dimensions['A'].width = 9 #픽셀 수 = 간격
    ws4.column_dimensions['B'].width = 25.4
    ws4.column_dimensions['C'].width = 84.1
    ws4.column_dimensions['D'].width = 28.4

    for i in first_row:
        cell_da = "{}{}".format(i,1)
        ws4[cell_da].border = boxA
        ws4[cell_da].fill = gray_fill
        ws4[cell_da].alignment = center_ali

        
    ws5['A1'] = '번호'
    ws5['B1'] = '탐지일'
    ws5['C1'] = 'URL'
    ws5['D1'] = '업체명'

    ws5.column_dimensions['A'].width = 9 #픽셀 수 = 간격
    ws5.column_dimensions['B'].width = 25.4
    ws5.column_dimensions['C'].width = 84.1
    ws5.column_dimensions['D'].width = 28.4

    for i in first_row:
        cell_da = "{}{}".format(i,1)
        ws5[cell_da].border = boxA
        ws5[cell_da].fill = gray_fill
        ws5[cell_da].alignment = center_ali
        

excel_name = "{}_사이버위협_대응_일일상황보고통계.xlsx".format(today.strftime('%y%m%d'))
wb.close()
wb.save(excel_name)
conn.close() 