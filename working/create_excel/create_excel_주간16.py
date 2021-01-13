# -*- coding utf-8 -*-
import openpyxl
from openpyxl.styles import colors
from openpyxl.styles import Font, Color, Border, Side, Alignment
from datetime import date, timedelta
from openpyxl.styles import PatternFill
import pymssql

#DB 연결부분
def dbConnection () :
    global conn
    try :
        conn = pymssql.connect(host="192.168.100.12", user='MCF', password='MCF_!@#$', database='MCF', charset='utf8')
        print("DB Connect Success!")
    except Exception as e:    
        print("DB Connect 실패!")


#conn = pymssql.connect(host="127.0.0.1", user='tester', password='1234', database='AdventureWorks2014')
dbConnection()
cursor = conn.cursor() 

def insert_value(ws = openpyxl.workbook.workbook.Workbook, sql_exe = None):
    cursor.execute(sql_exe)
    row = cursor.fetchone()
    count = 1
    while(row):
        count += 1
        
        cell_A = "A{}".format(count)
        ws[cell_A].border = boxA
        ws[cell_A].alignment = center_ali
        ws[cell_A] = count-1        
        row_index = 0
        
        for col in cell_range:
            cell_da = "{}{}".format(col,count)
            ws[cell_da].border = boxA            
            ws[cell_da].alignment = center_ali
            ws[cell_da] = row[row_index]
            row_index += 1
            if col is 'C':
                ws[cell_da].alignment = left_ali
        
        row = cursor.fetchone()        
    return count - 1
 

#cell style 정의
font_A = Font(name = '맑은 고딕', bold='True')
boxA = Border(top = Side(style='thin'), left = Side(style='thin'), right = Side(style='thin'),bottom = Side(style='thin'))
center_ali = Alignment(horizontal='center',vertical='center')
left_ali = Alignment(horizontal='general',vertical='center')
gray_fill = PatternFill('solid',fgColor='d8d8d8')


#날짜 변수 선언
today = date.today()
yesterday = (date.today() - timedelta(1))
week_day = (date.today() - timedelta(7))
day_check = today.weekday()

start_time = "05:00"
end_time = "16:00"


#범위
data_time_one = "{} {} ~ {} {}".format(yesterday.strftime('%m.%d'), end_time, today.strftime('%m.%d'), end_time)
data_time_two = "{} {} ~ {} {}".format(today.strftime('%m.%d'), start_time, today.strftime('%m.%d'), end_time)
data_time_week = f"{week_day.strftime('%m.%d')} {start_time} ~ {today.strftime('%m.%d')} {end_time}"


#시트 생성
wb = openpyxl.Workbook()
ws1 = wb.create_sheet('탐지건수', 0)
ws2 = wb.create_sheet('경유지(16시~16시)', 1)
ws3 = wb.create_sheet('유포지(16시~16시)', 2)
ws4 = wb.create_sheet('경유지(05시~16시)', 3)
ws5 = wb.create_sheet('유포지(05시~16시)', 4)
ws6 = wb['Sheet']
wb.remove(ws6)


#SQL 실행
excute_order ="SELECT b.create_dt '탐지일시', b.relay_url '경유지', (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END )  '업체명', (CASE (CASE left(C.WORK_TYPE_CD,4) WHEN 'KISA' then C.WORK_TYPE_CD else left(C.WORK_TYPE_CD,1)  end) when 'A' then '건강/의학' when 'B' then '게임' when 'C' then '교육/학원' when 'D' then '금융/부동산' when 'E' then '뉴스/미디어' when 'F' then '문학/예술' when 'G' then '비즈니스/경제' when 'H' then '사회/문화/종교' when 'I' then '생활/가정/취미' when 'J' then '서비스' when 'K' then '쇼핑' when 'L' then '스포츠/레저' when 'M' then '엔터테인먼트' when 'N' then '여행' when 'O' then '온라인교육' when 'P' then '유통/판매/운송' when 'Q' then '인터넷/컴퓨터' when 'R' then '정보통신/IT' when 'S' then '정치/행정' when 'T' then '제조' when 'U' then '커뮤니티' when 'V' then '학문' when 'K' then 'KISA' when 'KISA1' then '공공/국립' when 'KISA2' then '국가안보/주요 기반시설' when 'KISA3' then '군사이트(민간)' when 'KISA4' then '대북사이트' when 'KISA5' then '북한사이트' when 'KISA6' then '웹하드' when 'KISA7' then '해외' when 'KISA8' then '피싱' when 'KISA9' then '확인불가' else '' end)  '업종' FROM [MCF].[DBO].RELAY_URL_HISTORY B, [MCF].[DBO].DOMAIN_MASTER C WHERE	b.relay_domain = c.domain and b.create_dt between '{} 16:00:00' and '{} 16:00:00' and b.diffusion_url != 'http://ca-pub-5018740901612269' order by b.create_dt;".format(yesterday.strftime('%Y/%m/%d'), today.strftime('%Y/%m/%d'))
cursor.execute(excute_order)
row = cursor.fetchone()


#텍스트 파일 생성 및 excel 값 삽입
cell_range = ['B','C','D', 'E']
first_row = ['A','B','C','D', 'E']
first_row_one = ['A','B','C','D'] #탐지건수 시트용

count = 1
while(row):
    count += 1
    cell_A = "A{}".format(count)
    ws2[cell_A].border = boxA
    ws2[cell_A].alignment = center_ali
    ws2[cell_A] = count-1        
    row_index = 0
    
    for col in cell_range:
        cell_da = "{}{}".format(col,count)
        ws2[cell_da].border = boxA            
        ws2[cell_da].alignment = center_ali
        ws2[cell_da] = row[row_index]
        row_index += 1
        if col is 'C':
            ws2[cell_da].alignment = left_ali
    
    row = cursor.fetchone()        
count_relay_url_yes = count - 1


excute_order = "SELECT d.create_dt '탐지일시', d.diffusion_url '유포지', (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END ) '업체명', (CASE (CASE left(C.WORK_TYPE_CD,4) WHEN 'KISA' then C.WORK_TYPE_CD else left(C.WORK_TYPE_CD,1)  end) when 'A' then '건강/의학' when 'B' then '게임' when 'C' then '교육/학원' when 'D' then '금융/부동산' when 'E' then '뉴스/미디어' when 'F' then '문학/예술' when 'G' then '비즈니스/경제' when 'H' then '사회/문화/종교' when 'I' then '생활/가정/취미' when 'J' then '서비스' when 'K' then '쇼핑' when 'L' then '스포츠/레저' when 'M' then '엔터테인먼트' when 'N' then '여행' when 'O' then '온라인교육' when 'P' then '유통/판매/운송' when 'Q' then '인터넷/컴퓨터' when 'R' then '정보통신/IT' when 'S' then '정치/행정' when 'T' then '제조' when 'U' then '커뮤니티' when 'V' then '학문' when 'K' then 'KISA' when 'KISA1' then '공공/국립' when 'KISA2' then '국가안보/주요 기반시설' when 'KISA3' then '군사이트(민간)' when 'KISA4' then '대북사이트' when 'KISA5' then '북한사이트' when 'KISA6' then '웹하드' when 'KISA7' then '해외' when 'KISA8' then '피싱' when 'KISA9' then '확인불가' else '' end)  '업종' FROM [MCF].[DBO].DIFFUSION_URL_LIST B, [MCF].[DBO].DOMAIN_MASTER C, [MCF].[dbo].[DIFFUSION_ANALYSIS_GROUP_LIST] D WHERE b.diffusion_domain = c.domain and b.diffusion_url_hash = d.diffusion_url_hash and d.create_dt between  '{} 16:00:00' and '{} 16:00:00' order by d.create_dt;".format(yesterday.strftime('%Y/%m/%d') ,today.strftime('%Y/%m/%d'))
cursor.execute(excute_order)
row = cursor.fetchone()

count = 1
while(row):
    count += 1
    cell_A = "A{}".format(count)
    ws3[cell_A].border = boxA
    ws3[cell_A].alignment = center_ali
    ws3[cell_A] = count-1        
    row_index = 0
    
    for col in cell_range:
        cell_da = "{}{}".format(col,count)
        print(cell_da)
        ws3[cell_da].border = boxA            
        ws3[cell_da].alignment = center_ali
        ws3[cell_da] = row[row_index]
        row_index += 1
        if col is 'C':
            ws3[cell_da].alignment = left_ali
    
    row = cursor.fetchone()        
count_diffusion_url_yes = count - 1


excute_order ="SELECT b.create_dt '탐지일시', b.relay_url '경유지', (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END )  '업체명', (CASE (CASE left(C.WORK_TYPE_CD,4) WHEN 'KISA' then C.WORK_TYPE_CD else left(C.WORK_TYPE_CD,1)  end) when 'A' then '건강/의학' when 'B' then '게임' when 'C' then '교육/학원' when 'D' then '금융/부동산' when 'E' then '뉴스/미디어' when 'F' then '문학/예술' when 'G' then '비즈니스/경제' when 'H' then '사회/문화/종교' when 'I' then '생활/가정/취미' when 'J' then '서비스' when 'K' then '쇼핑' when 'L' then '스포츠/레저' when 'M' then '엔터테인먼트' when 'N' then '여행' when 'O' then '온라인교육' when 'P' then '유통/판매/운송' when 'Q' then '인터넷/컴퓨터' when 'R' then '정보통신/IT' when 'S' then '정치/행정' when 'T' then '제조' when 'U' then '커뮤니티' when 'V' then '학문' when 'K' then 'KISA' when 'KISA1' then '공공/국립' when 'KISA2' then '국가안보/주요 기반시설' when 'KISA3' then '군사이트(민간)' when 'KISA4' then '대북사이트' when 'KISA5' then '북한사이트' when 'KISA6' then '웹하드' when 'KISA7' then '해외' when 'KISA8' then '피싱' when 'KISA9' then '확인불가' else '' end)  '업종' FROM [MCF].[DBO].RELAY_URL_HISTORY B, [MCF].[DBO].DOMAIN_MASTER C WHERE	b.relay_domain = c.domain and b.create_dt between '{} 05:00:00' and '{} 16:00:00' and b.diffusion_url != 'http://ca-pub-5018740901612269' order by b.create_dt;".format(today.strftime('%Y/%m/%d'), today.strftime('%Y/%m/%d'))
cursor.execute(excute_order)
row = cursor.fetchone()

count = 1
while(row):
    count += 1
    cell_A = "A{}".format(count)
    ws4[cell_A].border = boxA
    ws4[cell_A].alignment = center_ali
    ws4[cell_A] = count-1        
    row_index = 0
    
    for col in cell_range:
        cell_da = "{}{}".format(col,count)
        ws4[cell_da].border = boxA            
        ws4[cell_da].alignment = center_ali
        ws4[cell_da] = row[row_index]
        row_index += 1
        if col is 'C':
            ws4[cell_da].alignment = left_ali
    
    row = cursor.fetchone()        
count_relay_url_to = count - 1


excute_order = "SELECT d.create_dt '탐지일시', d.diffusion_url '유포지', (CASE C.WORK_TYPE_CD WHEN 'KISA7' THEN '해외' ELSE c.company_nm END ) '업체명', (CASE (CASE left(C.WORK_TYPE_CD,4) WHEN 'KISA' then C.WORK_TYPE_CD else left(C.WORK_TYPE_CD,1)  end) when 'A' then '건강/의학' when 'B' then '게임' when 'C' then '교육/학원' when 'D' then '금융/부동산' when 'E' then '뉴스/미디어' when 'F' then '문학/예술' when 'G' then '비즈니스/경제' when 'H' then '사회/문화/종교' when 'I' then '생활/가정/취미' when 'J' then '서비스' when 'K' then '쇼핑' when 'L' then '스포츠/레저' when 'M' then '엔터테인먼트' when 'N' then '여행' when 'O' then '온라인교육' when 'P' then '유통/판매/운송' when 'Q' then '인터넷/컴퓨터' when 'R' then '정보통신/IT' when 'S' then '정치/행정' when 'T' then '제조' when 'U' then '커뮤니티' when 'V' then '학문' when 'K' then 'KISA' when 'KISA1' then '공공/국립' when 'KISA2' then '국가안보/주요 기반시설' when 'KISA3' then '군사이트(민간)' when 'KISA4' then '대북사이트' when 'KISA5' then '북한사이트' when 'KISA6' then '웹하드' when 'KISA7' then '해외' when 'KISA8' then '피싱' when 'KISA9' then '확인불가' else '' end)  '업종' FROM [MCF].[DBO].DIFFUSION_URL_LIST B, [MCF].[DBO].DOMAIN_MASTER C, [MCF].[dbo].[DIFFUSION_ANALYSIS_GROUP_LIST] D WHERE b.diffusion_domain = c.domain and b.diffusion_url_hash = d.diffusion_url_hash and d.create_dt between  '{} 05:00:00' and '{} 16:00:00' order by d.create_dt;".format(today.strftime('%Y/%m/%d') ,today.strftime('%Y/%m/%d'))
cursor.execute(excute_order)
row = cursor.fetchone()

count = 1
while(row):
    count += 1
    cell_A = "A{}".format(count)
    ws5[cell_A].border = boxA
    ws5[cell_A].alignment = center_ali
    ws5[cell_A] = count-1        
    row_index = 0
    
    for col in cell_range:
        cell_da = "{}{}".format(col,count)
        print(cell_da)
        ws5[cell_da].border = boxA            
        ws5[cell_da].alignment = center_ali
        ws5[cell_da] = row[row_index]
        row_index += 1
        if col is 'C':
            ws5[cell_da].alignment = left_ali
    
    row = cursor.fetchone()        
count_diffusion_url_to = count - 1


'''
###################
#####탐지 건수######
###################
'''

#A열
ws1['A2'] = '※ 탐지 건수'
ws1['A2'].font = font_A

ws1['A3'] = '일시'
ws1['A4'] =  data_time_one
ws1['A5'] =  data_time_two

ws1['A7'] = '※ 보고기준'
ws1['A7'].font = font_A

ws1['A8'] = ' - 16시 보고 : 전일 16시 ~ 금일 16시'
ws1['A9'] = '                   금일 05시 ~ 금일 16시'
ws1['A10'] = ' - 00시 보고 : 전일 00시 ~ 전일 24시'
ws1['A11'] = ' - 05시 보고 : 전일 05시 ~ 금일 05시'

ws1.column_dimensions['A'].width = 34.63 #픽셀 수 = 간격


#B열
ws1['B3'] = '경유지'
ws1['B4'] = count_relay_url_yes
ws1['B5'] = count_relay_url_to

ws1['C3'] = '유포지'
ws1['C4'] = count_diffusion_url_yes
ws1['C5'] = count_diffusion_url_to

ws1['D3'] = '합계'
ws1['D4'] = count_relay_url_yes + count_diffusion_url_yes
ws1['D5'] = count_relay_url_to + count_diffusion_url_to


for i in first_row_one:
    for x in range(3,6):
        cell_da = "{}{}".format(i,x)
        ws1[cell_da].border = boxA
        ws1[cell_da].alignment = center_ali
        

'''
#########################
####경유지(16시~16시)#####
#########################
'''

ws2['A1'] = '번호'
ws2['B1'] = '탐지일'
ws2['C1'] = 'URL'
ws2['D1'] = '업체명'
ws2['E1'] = '업종'

ws2.column_dimensions['A'].width = 9 #픽셀 수 = 간격
ws2.column_dimensions['B'].width = 25.4
ws2.column_dimensions['C'].width = 84.1
ws2.column_dimensions['D'].width = 28.4
ws2.column_dimensions['E'].width = 18

for i in first_row:
    cell_da = "{}{}".format(i,1)
    ws2[cell_da].border = boxA
    ws2[cell_da].fill = gray_fill
    ws2[cell_da].alignment = center_ali

'''
###########################
#####유포지(16시~16시)######
###########################
'''

ws3['A1'] = '번호'
ws3['B1'] = '탐지일'
ws3['C1'] = 'URL'
ws3['D1'] = '업체명'
ws3['E1'] = '업종'

ws3.column_dimensions['A'].width = 9 #픽셀 수 = 간격
ws3.column_dimensions['B'].width = 25.4
ws3.column_dimensions['C'].width = 84.1
ws3.column_dimensions['D'].width = 28.4
ws3.column_dimensions['E'].width = 18

for i in first_row:
    cell_da = "{}{}".format(i,1)
    ws3[cell_da].border = boxA
    ws3[cell_da].fill = gray_fill
    ws3[cell_da].alignment = center_ali

'''
#########################
####경유지(05시~16시)#####
#########################
'''

ws4['A1'] = '번호'
ws4['B1'] = '탐지일'
ws4['C1'] = 'URL'
ws4['D1'] = '업체명'
ws4['E1'] = '업종'

ws4.column_dimensions['A'].width = 9 #픽셀 수 = 간격
ws4.column_dimensions['B'].width = 25.4
ws4.column_dimensions['C'].width = 84.1
ws4.column_dimensions['D'].width = 28.4
ws4.column_dimensions['E'].width = 18

for i in first_row:
    cell_da = "{}{}".format(i,1)
    ws4[cell_da].border = boxA
    ws4[cell_da].fill = gray_fill
    ws4[cell_da].alignment = center_ali

'''
###########################
#####유포지(05시~16시)######
###########################
'''

ws5['A1'] = '번호'
ws5['B1'] = '탐지일'
ws5['C1'] = 'URL'
ws5['D1'] = '업체명'
ws5['E1'] = '업종'

ws5.column_dimensions['A'].width = 9 #픽셀 수 = 간격
ws5.column_dimensions['B'].width = 25.4
ws5.column_dimensions['C'].width = 84.1
ws5.column_dimensions['D'].width = 28.4
ws5.column_dimensions['E'].width = 18

for i in first_row:
    cell_da = "{}{}".format(i,1)
    ws5[cell_da].border = boxA
    ws5[cell_da].fill = gray_fill
    ws5[cell_da].alignment = center_ali
       

excel_name = "{}_사이버위협_대응_일일상황보고통계.xlsx".format(today.strftime('%y%m%d'))
wb.close()
wb.save(excel_name)
conn.close() 